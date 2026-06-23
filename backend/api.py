"""
FastAPI backend — serves World Bank procurement data to the React dashboard.
Run: uvicorn api:app --port 8080 --host 127.0.0.1
"""

import io
import csv
import os
import sys
import time
import subprocess
import threading
import json
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from fastapi import FastAPI, Query, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from dotenv import load_dotenv
import re
import requests
from bs4 import BeautifulSoup

from db import db, q, ensure_support_tables, get_app_settings_map
from services.tech import build_tech_notice_condition, classify_notice_tech, looks_like_tech_bidder
from services.contact_enrichment import search_company_contact
from services.gemini_service import classify_and_enrich_with_gemini
from services.bidder_extraction import (
    parse_notice_bidder_details, extract_bidders_list, extract_bidders_from_description,
    extract_awarded_bidders, fetch_bidders_from_detail_page, _infer_notice_category,
    _merge_categories, _clean_name, _dedupe,
)

# Load environment variables from .env file
load_dotenv()

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Target countries
DEFAULT_COUNTRIES = [
    "Zambia", "Sierra Leone", "Ethiopia", "Malawi", "Kenya",
    "Ghana", "Angola", "Central African Republic", "Guinea",
    "Gambia", "Botswana", "Benin", "Somalia, Federal Republic of", "Tanzania",
    "Mozambique", "Rwanda"
]


def get_country_fetch_status_rows():
    ensure_support_tables()
    rows = q("""
        WITH notice_counts AS (
            SELECT country,
                   COUNT(*) AS current_row_count,
                   MIN(notice_date)::text AS current_first_notice_date,
                   MAX(notice_date)::text AS current_last_notice_date,
                   MAX(fetched_at)::text AS current_last_fetched_at
            FROM procurement_notices
            WHERE country IS NOT NULL AND TRIM(country) <> ''
            GROUP BY country
        )
        SELECT
            c.name AS country,
            COALESCE(s.status,
                CASE WHEN COALESCE(n.current_row_count, 0) > 0 THEN 'success' ELSE 'not_started' END
            ) AS status,
            s.last_started_at::text,
            s.last_finished_at::text,
            s.last_success_at::text,
            s.last_attempted_since::text,
            s.last_page_size,
            COALESCE(s.fetched_records, 0) AS fetched_records,
            COALESCE(s.new_records, 0) AS new_records,
            COALESCE(s.total_available, 0) AS total_available,
            COALESCE(n.current_row_count, s.row_count, 0) AS row_count,
            COALESCE(n.current_first_notice_date, s.first_notice_date::text) AS first_notice_date,
            COALESCE(n.current_last_notice_date, s.last_notice_date::text) AS last_notice_date,
            n.current_last_fetched_at AS last_fetched_at,
            s.error_msg,
            s.api_url,
            COALESCE(s.retry_count, 0) AS retry_count,
            s.updated_at::text
        FROM target_countries c
        LEFT JOIN country_fetch_status s ON s.country = c.name
        LEFT JOIN notice_counts n ON n.country = c.name
        ORDER BY c.name
    """)

    def explain(row):
        if row["status"] == "failed":
            return row["error_msg"] or "Last fetch failed."
        if row["status"] == "running":
            return "Fetch is currently running."
        if row["row_count"]:
            if row["fetched_records"]:
                return "Fetched successfully in the last run."
            return "Has stored notices; no new notices were added in the last run."
        if row["status"] == "no_recent_notices":
            return "No notices matched the current incremental date window."
        if row["status"] == "no_data":
            return "Fetch completed, but no notices were stored for this country."
        return "No fetch has been recorded for this country yet."

    output = []
    for row in rows:
        item = dict(row)
        item["explanation"] = explain(item)
        output.append(item)
    return output

# ── Shared filter builder ─────────────────────────────────────────────────────

def build_where(country, notice_type, status, from_date, to_date, search, tech_only=False):
    filters = ["1=1"]
    params  = []

    if country:
        filters.append("country = %s")
        params.append(country)
    if notice_type:
        filters.append("""(
            notice_type = %s OR
            notice_type ILIKE %s OR
            notice_type ILIKE %s
        )""")
        full = {
            "IFB":  "%Invitation for Bids%",
            "REOI": "%Expression of Interest%",
            "Contract Award": "%Contract Award%",
            "Award": "%Contract Award%",
        }
        params.append(notice_type)
        params.append(full.get(notice_type, f"%{notice_type}%"))
        params.append(f"%{notice_type}%")
    if status:
        filters.append("status ILIKE %s")
        params.append(status)
    if from_date:
        filters.append("notice_date >= %s")
        params.append(from_date)
    if to_date:
        filters.append("notice_date <= %s")
        params.append(to_date)
    if search:
        filters.append("(title ILIKE %s OR description ILIKE %s OR project_name ILIKE %s)")
        like = f"%{search}%"
        params.extend([like, like, like])
    if tech_only:
        tech_filter, tech_params = build_tech_notice_condition()
        filters.append(tech_filter)
        params.extend(tech_params)

    return " AND ".join(filters), params


# ── Notices ───────────────────────────────────────────────────────────────────

@app.get("/api/notices")
def get_notices(
    country:     Optional[str]  = Query(None),
    notice_type: Optional[str]  = Query(None),
    status:      Optional[str]  = Query(None),
    from_date:   Optional[date] = Query(None),
    to_date:     Optional[date] = Query(None),
    search:      Optional[str]  = Query(None),
    tech_only:   bool           = Query(False),
    page:        int            = Query(1, ge=1),
    page_size:   int            = Query(25, le=100),
):
    where, params = build_where(country, notice_type, status, from_date, to_date, search, tech_only)
    offset = (page - 1) * page_size

    country_rows = q("SELECT name FROM target_countries ORDER BY name")
    if country_rows:
        available_countries = [r["name"] for r in country_rows]
    else:
        available_countries = [r["country"] for r in q("""
            SELECT DISTINCT country FROM procurement_notices
            WHERE country IS NOT NULL AND TRIM(country) <> ''
            ORDER BY country
        """)]

    rows = q(
        f"""SELECT
                id, project_id, project_name, country, notice_type,
                procurement_method, title, description,
                COALESCE(submission_date, submission_deadline::date)::text AS submission_date,
                notice_date::text,
                contract_amount, currency, borrower, contact_email, url, status,
                fetched_at
            FROM procurement_notices
            WHERE {where}
            ORDER BY notice_date DESC NULLS LAST
            LIMIT %s OFFSET %s""",
        params + [page_size, offset]
    )
    total = q(f"SELECT COUNT(*) as cnt FROM procurement_notices WHERE {where}", params)

    output_rows = []
    for row in rows:
        item = dict(row)
        item.update(classify_notice_tech(item))
        output_rows.append(item)

    return {
        "total":               total[0]["cnt"],
        "page":                page,
        "page_size":           page_size,
        "available_countries": available_countries,
        "data":                output_rows,
    }


# ── Stats ─────────────────────────────────────────────────────────────────────

@app.get("/api/stats")
def get_stats(country: Optional[str] = Query(None)):
    try:
        selected_country = country.strip() if country else None

        general_total_count = q("SELECT COUNT(*) as total FROM procurement_notices")
        general_total_ifb = q("SELECT COUNT(*) as total FROM procurement_notices WHERE notice_type = 'IFB'")
        general_total_reoi = q("SELECT COUNT(*) as total FROM procurement_notices WHERE notice_type = 'REOI'")
        general_total_award = q("""
            SELECT COUNT(*) as total
            FROM procurement_notices
            WHERE notice_type = 'Contract Award' OR notice_type = 'Award'
        """)
        general_countries_count = q("""
            SELECT COUNT(DISTINCT country) as count
            FROM procurement_notices
            WHERE country IS NOT NULL
        """)
        last_fetched = q("""
            SELECT MAX(fetched_at) as last_fetched
            FROM procurement_notices
        """)

        available_countries = q("""
            SELECT DISTINCT country
            FROM procurement_notices
            WHERE country IS NOT NULL AND TRIM(country) <> ''
            ORDER BY country
        """)
        configured_countries = q("""
            SELECT name
            FROM target_countries
            ORDER BY name
        """)

        configured_country_names = [r["name"] for r in configured_countries] if configured_countries else []
        active_country_names = [r["country"] for r in available_countries] if available_countries else []
        inactive_country_names = [name for name in configured_country_names if name not in set(active_country_names)]

        country_filter = ""
        params = []
        if selected_country:
            country_filter = " AND country = %s"
            params.append(selected_country)

        summary_total = q(f"SELECT COUNT(*) as total FROM procurement_notices WHERE 1=1{country_filter}", params)
        summary_ifb = q(f"SELECT COUNT(*) as total FROM procurement_notices WHERE notice_type = 'IFB'{country_filter}", params)
        summary_reoi = q(f"SELECT COUNT(*) as total FROM procurement_notices WHERE notice_type = 'REOI'{country_filter}", params)
        summary_award = q(f"""
            SELECT COUNT(*) as total
            FROM procurement_notices
            WHERE (notice_type = 'Contract Award' OR notice_type = 'Award'){country_filter}
        """, params)
        summary_statuses = q(f"""
            SELECT COUNT(DISTINCT status) as count
            FROM procurement_notices
            WHERE status IS NOT NULL AND TRIM(status) <> ''{country_filter}
        """, params)
        summary_borrowers = q(f"""
            SELECT COUNT(DISTINCT borrower) as count
            FROM procurement_notices
            WHERE borrower IS NOT NULL AND TRIM(borrower) <> ''{country_filter}
        """, params)

        by_country = q("""
            SELECT
                country,
                COUNT(*) as total_count,
                COUNT(CASE WHEN notice_type = 'IFB' THEN 1 END) as ifb_count,
                COUNT(CASE WHEN notice_type = 'REOI' THEN 1 END) as reoi_count,
                COUNT(CASE WHEN notice_type = 'Contract Award' OR notice_type = 'Award' THEN 1 END) as award_count
            FROM procurement_notices
            WHERE country IS NOT NULL
            GROUP BY country
            ORDER BY total_count DESC
            LIMIT 10
        """)

        by_month = q(f"""
            SELECT
                TO_CHAR(notice_date::date, 'YYYY-MM') as month,
                COUNT(*) as count
            FROM procurement_notices
            WHERE notice_date IS NOT NULL{country_filter}
            GROUP BY TO_CHAR(notice_date::date, 'YYYY-MM')
            ORDER BY month DESC
            LIMIT 12
        """, params)

        by_type = q(f"""
            SELECT notice_type, COUNT(*) as count
            FROM procurement_notices
            WHERE notice_type IS NOT NULL{country_filter}
            GROUP BY notice_type
            ORDER BY count DESC
        """, params)

        by_status = q(f"""
            SELECT status, COUNT(*) as count
            FROM procurement_notices
            WHERE status IS NOT NULL AND TRIM(status) <> ''{country_filter}
            GROUP BY status
            ORDER BY count DESC
        """, params)

        top_borrowers = q(f"""
            SELECT borrower, COUNT(*) as count
            FROM procurement_notices
            WHERE borrower IS NOT NULL AND TRIM(borrower) <> ''{country_filter}
            GROUP BY borrower
            ORDER BY count DESC, borrower ASC
            LIMIT 8
        """, params)

        recent = q(f"""
            SELECT id, country, title, project_id, notice_type, notice_date::text, url
            FROM procurement_notices
            WHERE 1=1{country_filter}
            ORDER BY fetched_at DESC
            LIMIT 5
        """, params)

        return {
            "selected_country": selected_country,
            "available_countries": active_country_names,
            "countries_overview": {
                "configured": configured_country_names,
                "active": active_country_names,
                "inactive": inactive_country_names,
                "configured_count": len(configured_country_names),
                "active_count": len(active_country_names),
                "inactive_count": len(inactive_country_names),
            },
            "summary": {
                "total": general_total_count[0]["total"] if general_total_count else 0,
                "total_ifb": general_total_ifb[0]["total"] if general_total_ifb else 0,
                "total_reoi": general_total_reoi[0]["total"] if general_total_reoi else 0,
                "total_award": general_total_award[0]["total"] if general_total_award else 0,
                "countries": general_countries_count[0]["count"] if general_countries_count else 0,
                "last_fetched": last_fetched[0]["last_fetched"] if last_fetched else None
            },
            "country_summary": {
                "country": selected_country,
                "total": summary_total[0]["total"] if summary_total else 0,
                "total_ifb": summary_ifb[0]["total"] if summary_ifb else 0,
                "total_reoi": summary_reoi[0]["total"] if summary_reoi else 0,
                "total_award": summary_award[0]["total"] if summary_award else 0,
                "statuses": summary_statuses[0]["count"] if summary_statuses else 0,
                "borrowers": summary_borrowers[0]["count"] if summary_borrowers else 0,
            },
            "by_country": [dict(r) for r in by_country] if by_country else [],
            "by_month": [dict(r) for r in by_month] if by_month else [],
            "by_type": [dict(r) for r in by_type] if by_type else [],
            "by_status": [dict(r) for r in by_status] if by_status else [],
            "top_borrowers": [dict(r) for r in top_borrowers] if top_borrowers else [],
            "recent": [dict(r) for r in recent] if recent else []
        }

    except Exception as e:
        return {
            "summary": {"total": 0, "countries": 0},
            "country_summary": {"country": selected_country if 'selected_country' in locals() else None, "total": 0, "total_ifb": 0, "total_reoi": 0, "total_award": 0, "statuses": 0, "borrowers": 0},
            "available_countries": [],
            "countries_overview": {"configured": [], "active": [], "inactive": [], "configured_count": 0, "active_count": 0, "inactive_count": 0},
            "selected_country": selected_country if 'selected_country' in locals() else None,
            "by_country": [],
            "recent": [],
            "by_month": [],
            "by_type": [],
            "by_status": [],
            "top_borrowers": [],
            "error": str(e)
        }


@app.get("/api/dashboard")
def get_dashboard(
    country: Optional[str] = Query(None),
    notice_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
):
    try:
        ensure_support_tables()
        where, params = build_where(country, notice_type, status, from_date, to_date, search)
        settings = get_app_settings_map()

        configured_rows = q("SELECT name FROM target_countries ORDER BY name")
        configured = [row["name"] for row in configured_rows]
        active_rows = q("""
            SELECT DISTINCT country
            FROM procurement_notices
            WHERE country IS NOT NULL AND TRIM(country) <> ''
            ORDER BY country
        """)
        active = [row["country"] for row in active_rows]
        active_set = set(active)
        inactive = [name for name in configured if name not in active_set]

        summary = q(f"""
            SELECT
                COUNT(*) AS total,
                COUNT(CASE WHEN notice_type = 'IFB' THEN 1 END) AS total_ifb,
                COUNT(CASE WHEN notice_type = 'REOI' THEN 1 END) AS total_reoi,
                COUNT(CASE WHEN notice_type = 'Contract Award' OR notice_type = 'Award' THEN 1 END) AS total_award,
                COUNT(DISTINCT country) AS countries,
                COUNT(DISTINCT borrower) FILTER (WHERE borrower IS NOT NULL AND TRIM(borrower) <> '') AS borrowers,
                COUNT(DISTINCT status) FILTER (WHERE status IS NOT NULL AND TRIM(status) <> '') AS statuses,
                MAX(fetched_at) AS last_fetched
            FROM procurement_notices
            WHERE {where}
        """, params)[0]

        by_country = q(f"""
            SELECT
                country,
                COUNT(*) AS total_count,
                COUNT(CASE WHEN notice_type = 'IFB' THEN 1 END) AS ifb_count,
                COUNT(CASE WHEN notice_type = 'REOI' THEN 1 END) AS reoi_count,
                COUNT(CASE WHEN notice_type = 'Contract Award' OR notice_type = 'Award' THEN 1 END) AS award_count
            FROM procurement_notices
            WHERE {where} AND country IS NOT NULL AND TRIM(country) <> ''
            GROUP BY country
            ORDER BY total_count DESC, country ASC
            LIMIT 12
        """, params)

        by_type = q(f"""
            SELECT notice_type, COUNT(*) AS count
            FROM procurement_notices
            WHERE {where} AND notice_type IS NOT NULL AND TRIM(notice_type) <> ''
            GROUP BY notice_type
            ORDER BY count DESC, notice_type ASC
        """, params)

        by_status = q(f"""
            SELECT status, COUNT(*) AS count
            FROM procurement_notices
            WHERE {where} AND status IS NOT NULL AND TRIM(status) <> ''
            GROUP BY status
            ORDER BY count DESC, status ASC
        """, params)

        by_month = q(f"""
            SELECT TO_CHAR(notice_date::date, 'YYYY-MM') AS month, COUNT(*) AS count
            FROM procurement_notices
            WHERE {where} AND notice_date IS NOT NULL
            GROUP BY TO_CHAR(notice_date::date, 'YYYY-MM')
            ORDER BY month DESC
            LIMIT 12
        """, params)

        top_borrowers = q(f"""
            SELECT borrower, COUNT(*) AS count, MAX(notice_date)::text AS last_notice_date
            FROM procurement_notices
            WHERE {where} AND borrower IS NOT NULL AND TRIM(borrower) <> ''
            GROUP BY borrower
            ORDER BY count DESC, borrower ASC
            LIMIT 10
        """, params)

        recent = q(f"""
            SELECT id, country, title, project_id, notice_type, notice_date::text, url, borrower, status
            FROM procurement_notices
            WHERE {where}
            ORDER BY fetched_at DESC
            LIMIT 8
        """, params)

        data_quality = q(f"""
            SELECT
                COUNT(*) FILTER (WHERE borrower IS NULL OR TRIM(borrower) = '') AS missing_borrower,
                COUNT(*) FILTER (WHERE contact_email IS NULL OR TRIM(contact_email) = '') AS missing_contact_email,
                COUNT(*) FILTER (WHERE submission_date IS NULL) AS missing_submission_date,
                COUNT(*) FILTER (WHERE procurement_method IS NULL OR TRIM(procurement_method) = '') AS missing_procurement_method
            FROM procurement_notices
            WHERE {where}
        """, params)[0]

        deadlines = q(f"""
            SELECT
                COUNT(*) FILTER (WHERE COALESCE(submission_date, submission_deadline::date) BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '7 days') AS upcoming_7_days,
                COUNT(*) FILTER (WHERE COALESCE(submission_date, submission_deadline::date) BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days') AS upcoming_30_days,
                COUNT(*) FILTER (WHERE COALESCE(submission_date, submission_deadline::date) < CURRENT_DATE AND status ILIKE 'Active') AS overdue_active
            FROM procurement_notices
            WHERE {where}
        """, params)[0]

        upcoming_deadlines = q(f"""
            SELECT country, title, project_id, COALESCE(submission_date, submission_deadline::date)::text AS submission_date, notice_type, url
            FROM procurement_notices
            WHERE {where}
              AND COALESCE(submission_date, submission_deadline::date) IS NOT NULL
              AND COALESCE(submission_date, submission_deadline::date) BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days'
            ORDER BY COALESCE(submission_date, submission_deadline::date) ASC
            LIMIT 10
        """, params)

        recent_changes = q(f"""
            SELECT country, title, project_id, notice_type, status, updated_at::text, notice_date::text
            FROM procurement_notices
            WHERE {where}
              AND updated_at IS NOT NULL
            ORDER BY updated_at DESC
            LIMIT 10
        """, params)

        activity_scores = q("""
            SELECT
                c.name AS country,
                COUNT(p.id) FILTER (WHERE p.notice_date >= CURRENT_DATE - INTERVAL '30 days') AS recent_count
            FROM target_countries c
            LEFT JOIN procurement_notices p ON p.country = c.name
            GROUP BY c.name
            ORDER BY c.name
        """)
        scored = []
        for row in activity_scores:
            recent_count = int(row["recent_count"] or 0)
            if recent_count >= 15:
                score = "High"
            elif recent_count >= 5:
                score = "Moderate"
            elif recent_count >= 1:
                score = "Low"
            else:
                score = "Inactive"
            scored.append({"country": row["country"], "recent_count": recent_count, "score": score})

        fetch_runs = q("""
            SELECT run_at::text, country, fetched, new_records, success, error_msg
            FROM fetch_runs
            ORDER BY run_at DESC
            LIMIT 8
        """)

        fetch_health = {
            "running": _fetch_status["running"],
            "last_triggered": _fetch_status["last_triggered"],
            "last_finished": _fetch_status["last_finished"],
            "last_result": _fetch_status["last_result"],
            "recent_runs": [dict(r) for r in fetch_runs],
            "country_statuses": get_country_fetch_status_rows(),
        }

        return {
            "filters": {
                "country": country,
                "notice_type": notice_type,
                "status": status,
                "from_date": str(from_date) if from_date else "",
                "to_date": str(to_date) if to_date else "",
                "search": search or "",
            },
            "settings": {
                "baseline_date": settings.get("baseline_date", "2025-01-01"),
                "country_batch": int(float(settings.get("country_batch", 5))),
                "request_delay": float(settings.get("request_delay", 1.2)),
                "auto_sync_hour": settings.get("auto_sync_hour", "06:00"),
            },
            "available_countries": active,
            "countries_overview": {
                "configured": configured,
                "active": active,
                "inactive": inactive,
                "configured_count": len(configured),
                "active_count": len(active),
                "inactive_count": len(inactive),
            },
            "summary": dict(summary),
            "by_country": [dict(r) for r in by_country],
            "by_type": [dict(r) for r in by_type],
            "by_status": [dict(r) for r in by_status],
            "by_month": [dict(r) for r in by_month],
            "top_borrowers": [dict(r) for r in top_borrowers],
            "recent": [dict(r) for r in recent],
            "data_quality": dict(data_quality),
            "deadlines": {
                **dict(deadlines),
                "upcoming": [dict(r) for r in upcoming_deadlines],
            },
            "recent_changes": [dict(r) for r in recent_changes],
            "activity_scores": scored,
            "fetch_health": fetch_health,
        }
    except Exception as e:
        return {
            "summary": {"total": 0, "countries": 0},
            "available_countries": [],
            "countries_overview": {"configured": [], "active": [], "inactive": [], "configured_count": 0, "active_count": 0, "inactive_count": 0},
            "by_country": [],
            "by_type": [],
            "by_status": [],
            "by_month": [],
            "top_borrowers": [],
            "recent": [],
            "data_quality": {},
            "deadlines": {"upcoming": []},
            "recent_changes": [],
            "activity_scores": [],
            "fetch_health": {"recent_runs": [], "country_statuses": []},
            "error": str(e)
        }


# ── Country Management ────────────────────────────────────────────────────────

class CountryBody(BaseModel):
    name: str


class GeneralSettingsBody(BaseModel):
    baseline_date: Optional[str] = None
    country_batch: Optional[int] = None
    request_delay: Optional[float] = None
    auto_sync_hour: Optional[str] = None


@app.get("/api/settings/countries")
def list_countries():
    rows = q("SELECT name, added_at FROM target_countries ORDER BY name")
    return [dict(r) for r in rows]


@app.post("/api/settings/countries", status_code=201)
def add_country(body: CountryBody):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Country name cannot be empty")
    try:
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO target_countries (name) VALUES (%s) ON CONFLICT (name) DO NOTHING RETURNING name",
                    (name,)
                )
                result = cur.fetchone()
            conn.commit()
        if not result:
            raise HTTPException(409, f"'{name}' already exists")
        return {"name": name, "added": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/settings/countries/{name}")
def remove_country(name: str):
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM target_countries WHERE name = %s RETURNING name", (name,))
            deleted = cur.fetchone()
        conn.commit()
    if not deleted:
        raise HTTPException(404, f"'{name}' not found")
    return {"name": name, "deleted": True}


@app.get("/api/settings/general")
def get_general_settings():
    settings = get_app_settings_map()
    return {
        "baseline_date": settings.get("baseline_date", "2025-01-01"),
        "country_batch": int(float(settings.get("country_batch", 5))),
        "request_delay": float(settings.get("request_delay", 1.2)),
        "auto_sync_hour": settings.get("auto_sync_hour", "06:00"),
        "updated_at": settings.get("updated_at"),
    }


@app.put("/api/settings/general")
def update_general_settings(body: GeneralSettingsBody):
    ensure_support_tables()
    updates = {
        "baseline_date": body.baseline_date,
        "country_batch": str(body.country_batch) if body.country_batch is not None else None,
        "request_delay": str(body.request_delay) if body.request_delay is not None else None,
        "auto_sync_hour": body.auto_sync_hour,
    }

    with db() as conn:
        with conn.cursor() as cur:
            for key, value in updates.items():
                if value is None:
                    continue
                cur.execute(
                    """
                    INSERT INTO app_settings (key, value, updated_at)
                    VALUES (%s, %s, NOW())
                    ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
                    """,
                    (key, value),
                )
        conn.commit()

    return get_general_settings()


# ── Fetch trigger ─────────────────────────────────────────────────────────────

_fetch_status = {
    "running":        False,
    "last_triggered": None,
    "last_finished":  None,
    "last_result":    None,
}


@app.post("/api/fetch")
def trigger_fetch():
    if _fetch_status["running"]:
        return {"status": "already_running", "message": "Fetch already in progress"}

    fetcher_path = os.path.join(os.path.dirname(__file__), "fetcher_once.py")
    if not os.path.exists(fetcher_path):
        raise HTTPException(404, "fetcher_once.py not found next to api.py")

    _fetch_status["running"]        = True
    _fetch_status["last_triggered"] = datetime.utcnow().isoformat()
    _fetch_status["last_finished"]  = None
    _fetch_status["last_result"]    = None

    def run_fetcher():
        try:
            result = subprocess.run(
                [sys.executable, fetcher_path],
                capture_output=True, text=True, timeout=18000
            )
            _fetch_status["last_result"] = {
                "exit_code": result.returncode,
                "stdout":    result.stdout[-3000:],
                "stderr":    result.stderr[-1000:],
                "success":   result.returncode == 0,
            }
        except subprocess.TimeoutExpired:
            _fetch_status["last_result"] = {
                "exit_code": -1, "stdout": "",
                "stderr": "Fetcher timed out after 5 hours", "success": False,
            }
        except Exception as e:
            _fetch_status["last_result"] = {
                "exit_code": -1, "stdout": "", "stderr": str(e), "success": False,
            }
        finally:
            _fetch_status["running"]       = False
            _fetch_status["last_finished"] = datetime.utcnow().isoformat()

    threading.Thread(target=run_fetcher, daemon=True).start()
    return {"status": "started", "message": "Fetcher started in background"}


@app.get("/api/fetch/status")
def get_fetch_status():
    return {
        "running": _fetch_status["running"],
        "last_triggered": _fetch_status["last_triggered"],
        "last_finished": _fetch_status["last_finished"],
        "last_result": _fetch_status["last_result"],
    }





# ── Excel Export ────────────────────────────────────────────────────────────────

EXPORT_FIELD_CONFIG = {
    'country': ("Country", "country", 15),
    'notice_type': ("Notice Type", "notice_type", 16),
    'status': ("Status", "status", 12),
    'title': ("Opportunity Title", "title", 50),
    'project_id': ("Project ID", "project_id", 18),
    'project_name': ("Project Name", "project_name", 40),
    'procurement_method': ("Procurement Method", "procurement_method", 25),
    'borrower': ("Host Institution", "borrower", 35),
    'notice_date': ("Notice Date", "notice_date", 15),
    'awarded_date': ("Awarded Date", "awarded_date", 15),
    'submission_date': ("Submission Deadline", "submission_date", 20),
    'contract_amount': ("Contract Amount", "contract_amount", 20),
    'currency': ("Currency", "currency", 10),
    'contact_email': ("Contact Email", "contact_email", 30),
    'url': ("World Bank Link", "url", 25),
    'bidders': ("Bidders", "bidders", 40),
    'is_tech': ("Tech Opportunity", "is_tech", 16),
    'tech_category': ("Tech Category", "tech_category", 28),
    'overview': ("Overview", "overview", 60),
    'requirements': ("Requirements", "requirements", 60),
    'description': ("Description", "description", 60),
    'fetched_at': ("Fetched Date", "fetched_at", 18),
    'opportunity_id': ("Opportunity ID", "opportunity_id", 20),
}

CUSTOM_BIDDER_EXPORT_FIELD_CONFIG = {
    'country': ("Country", "country", 15),
    'notice_type': ("Notice Type", "notice_type", 16),
    'status': ("Status", "status", 12),
    'title': ("Opportunity Title", "title", 50),
    'project_id': ("Project ID", "project_id", 18),
    'project_name': ("Project Name", "project_name", 40),
    'procurement_method': ("Procurement Method", "procurement_method", 25),
    'borrower': ("Host Institution", "borrower", 35),
    'notice_date': ("Notice Date", "notice_date", 15),
    'submission_date': ("Submission Deadline", "submission_date", 20),
    'is_tech': ("Tech Opportunity", "is_tech", 16),
    'tech_category': ("Tech Category", "tech_category", 28),
    'overview': ("Overview", "overview", 60),
    'requirements': ("Requirements", "requirements", 60),
    'bidder_name': ("Bidder Name", "bidder_name", 40),
    'bidder_country': ("Bidder Country", "bidder_country", 20),
    'bidder_status': ("Bidder Status", "bidder_status", 18),
    'won': ("Won", "won", 10),
    'bid_price_at_opening': ("Bid Price at Opening", "bid_price_at_opening", 20),
    'opening_currency': ("Opening Bid Currency", "opening_currency", 16),
    'evaluated_bid_price': ("Evaluated Bid Price", "evaluated_bid_price", 20),
    'evaluated_bid_currency': ("Evaluated Bid Currency", "evaluated_bid_currency", 18),
    'final_evaluation_price': ("Final Evaluation Price", "final_evaluation_price", 20),
    'final_evaluation_currency': ("Final Evaluation Currency", "final_evaluation_currency", 18),
    'winner_contract_amount': ("Contract Amount (Winner Only)", "winner_contract_amount", 24),
    'winner_contract_currency': ("Contract Currency", "winner_contract_currency", 16),
    'contact_email': ("Contact Email", "contact_email", 30),
    'url': ("World Bank Link", "url", 25),
    'opportunity_id': ("Opportunity ID", "opportunity_id", 20),
}

EXPORT_FIELD_MAPPING = {
    'country': 'country',
    'notice_type': 'notice_type',
    'status': 'status',
    'title': 'title',
    'project_id': 'project_id',
    'project_name': 'project_name',
    'procurement_method': 'procurement_method',
    'borrower': 'borrower',
    'notice_date': 'notice_date',
    'awarded_date': '(SELECT MAX(ba.award_date) FROM bidder_awards ba WHERE ba.notice_id = procurement_notices.id) AS awarded_date',
    'submission_date': 'COALESCE(submission_date, submission_deadline::date)::text AS submission_date',
    'contract_amount': 'contract_amount',
    'currency': 'currency',
    'contact_email': 'contact_email',
    'url': 'url',
    'description': 'description',
    'fetched_at': 'fetched_at',
    'opportunity_id': 'id as opportunity_id',
    'bidders': 'description',
    'overview': 'description',
    'requirements': 'description',
    'is_tech': 'description',
    'tech_category': 'description',
}

DEFAULT_EXPORT_FIELDS = [
    'country', 'notice_type', 'status', 'title', 'project_id', 'project_name',
    'procurement_method', 'borrower', 'notice_date', 'awarded_date', 'submission_date',
    'contract_amount', 'currency', 'contact_email', 'url'
]


def resolve_export_fields(fields: Optional[str]) -> List[str]:
    selected_fields = [f.strip() for f in (fields or "").split(',') if f.strip()]
    if not selected_fields:
        return DEFAULT_EXPORT_FIELDS.copy()
    return [field for field in selected_fields if field in EXPORT_FIELD_CONFIG]


DEFAULT_CUSTOM_BIDDER_EXPORT_FIELDS = [
    'country', 'notice_type', 'title', 'project_id', 'project_name',
    'borrower', 'notice_date', 'bidder_name', 'bidder_country',
    'bidder_status', 'won', 'evaluated_bid_price', 'evaluated_bid_currency',
    'winner_contract_amount', 'winner_contract_currency', 'url'
]

BASE_TEMPLATE_EXPORT_COLUMNS = [
    ("CONTRACTOR", "contractor", 20.6640625),
    ("CONTRACTOR'S COUNTRY", "contractor_country", 35.77734375),
    ("TENDER", "tender", 19.88671875),
    ("AWARD DATE", "award_date", 31.5546875),
    ("AWARD AMOUNT", "award_amount", 41.109375),
    ("CONTRACTOR REFERENCE NUMBER", "contractor_reference_number", 59.6640625),
    ("CONTACT", "contact", 31.109375),
]


def get_template_export_columns(include_deadline: bool = False):
    columns = BASE_TEMPLATE_EXPORT_COLUMNS.copy()
    if include_deadline:
        columns.insert(4, ("DEADLINE DATE", "deadline_date", 20))
    return columns + [(None, f"blank_{i}", 13) for i in range(len(columns) + 1, 26)]


def resolve_custom_bidder_export_fields(fields: Optional[str]) -> List[str]:
    selected_fields = [f.strip() for f in (fields or "").split(',') if f.strip()]
    if not selected_fields:
        return DEFAULT_CUSTOM_BIDDER_EXPORT_FIELDS.copy()
    return [field for field in selected_fields if field in CUSTOM_BIDDER_EXPORT_FIELD_CONFIG]


def _clean_export_text(text: Optional[str]) -> str:
    if not text:
        return ""
    cleaned = BeautifulSoup(str(text), "html.parser").get_text(" ")
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()


def extract_notice_overview(text: Optional[str]) -> Optional[str]:
    cleaned = _clean_export_text(text)
    if not cleaned:
        return None

    overview_patterns = [
        r'(?:Scope of Contract|Scope of Work|Description|Assignment Title|Project Description)\s*:?\s*(.+?)(?=\s+(?:Loan/Credit/TF Info|Bid/Contract Reference No|Procurement Method|Awarded Bidder|Evaluated Bidder|Qualification|Eligibility|Requirements?)\s*:|$)',
        r'(?:The objective of|The project (?:will|is)|This assignment (?:will|is)|The consulting services include)\s+(.+?)(?=\s+(?:Qualification|Eligibility|Requirements?|Interested consultants|The attention of)\b|$)',
    ]
    for pattern in overview_patterns:
        match = re.search(pattern, cleaned, flags=re.IGNORECASE)
        if match:
            value = match.group(1).strip(" :-")
            if value:
                return value[:1000]
    return cleaned[:1000]


def extract_notice_requirements(text: Optional[str]) -> Optional[str]:
    cleaned = _clean_export_text(text)
    if not cleaned:
        return None

    section_patterns = [
        r'((?:Qualification|Eligibility|Requirements?|Minimum Qualification|Selection Criteria|Evaluation Criteria|Shortlisting Criteria|Interested consultants)[^:]{0,80}:?\s+.+?)(?=\s+(?:Submission|Deadline|Expressions of Interest|Further information|Contact|Awarded Bidder|Evaluated Bidder|Scope of Contract)\b|$)',
        r'((?:The attention of interested Consultants|Interested consultants should provide|Consultants may associate|A Consultant will be selected).+?)(?=\s+(?:Further information|Expressions of Interest|Submission|Deadline|Contact)\b|$)',
    ]
    sections = []
    for pattern in section_patterns:
        for match in re.finditer(pattern, cleaned, flags=re.IGNORECASE):
            value = match.group(1).strip(" :-")
            if value and value.lower() not in {item.lower() for item in sections}:
                sections.append(value[:1000])
    if sections:
        return " | ".join(sections)[:1500]

    hints = []
    keyword_hints = [
        ("qualification", "Professional qualifications or certifications"),
        ("experience", "Relevant experience on similar assignments"),
        ("financial", "Financial capacity"),
        ("technical", "Technical capability"),
        ("equipment", "Required equipment or tools"),
        ("license", "Valid licenses or registration"),
    ]
    lower = cleaned.lower()
    for keyword, label in keyword_hints:
        if keyword in lower:
            hints.append(label)
    return "; ".join(hints) if hints else None


def fetch_export_rows(where: str, params, selected_fields: List[str]):
    select_fields = []
    for field in selected_fields:
        mapped = EXPORT_FIELD_MAPPING.get(field)
        if mapped:
            select_fields.append(mapped)
    if 'bidders' in selected_fields and 'description' not in select_fields:
        select_fields.append('description')
    if any(field in selected_fields for field in ("is_tech", "tech_category")):
        for extra in ("title", "project_name", "procurement_method", "borrower_bid_reference", "description"):
            if extra not in select_fields:
                select_fields.append(extra)
    if not select_fields:
        select_fields = ['country', 'title', 'notice_date']

    seen = set()
    ordered_select = []
    for field in select_fields:
        if field not in seen:
            seen.add(field)
            ordered_select.append(field)

    return q(f"""
        SELECT {', '.join(ordered_select)}
        FROM procurement_notices
        WHERE {where}
        ORDER BY notice_date DESC
    """, params or ())


def fetch_custom_bidder_export_rows(where: str, params, selected_fields: List[str]):
    notice_rows = q(f"""
        SELECT
            country,
            notice_type,
            status,
            title,
            project_id,
            project_name,
            procurement_method,
            borrower,
            borrower_bid_reference,
            notice_no,
            notice_date::text,
            COALESCE(submission_date, submission_deadline::date)::text AS submission_date,
            contract_amount,
            currency,
            contact_email,
            url,
            id AS opportunity_id,
            description
        FROM procurement_notices
        WHERE {where}
        ORDER BY notice_date DESC
    """, params or ())

    bidder_rows: List[Dict[str, Any]] = []
    base_notice_fields = (
        "country", "notice_type", "status", "title", "project_id", "project_name",
        "procurement_method", "borrower", "notice_date", "submission_date",
        "contact_email", "url", "opportunity_id", "borrower_bid_reference", "notice_no", "description"
    )

    for notice in notice_rows:
        desc = notice.get("description") or ""
        bidder_details = parse_notice_bidder_details(desc)

        if bidder_details:
            for detail in bidder_details:
                won = detail.get("section") == "awarded"
                row = {field: notice.get(field) for field in base_notice_fields}
                row.update({
                    "bidder_name": detail.get("name"),
                    "bidder_country": detail.get("country"),
                    "bidder_status": detail.get("role"),
                    "won": "Yes" if won else "No",
                    "bid_price_at_opening": detail.get("opening_amount"),
                    "opening_currency": detail.get("opening_currency"),
                    "evaluated_bid_price": detail.get("evaluated_amount"),
                    "evaluated_bid_currency": detail.get("evaluated_currency"),
                    "final_evaluation_price": detail.get("final_amount"),
                    "final_evaluation_currency": detail.get("final_currency"),
                    "winner_contract_amount": (
                        detail.get("signed_amount")
                        if won and detail.get("signed_amount") is not None
                        else notice.get("contract_amount") if won else None
                    ),
                    "winner_contract_currency": (
                        (detail.get("signed_currency") or detail.get("currency") or notice.get("currency"))
                        if won else None
                    ),
                })
                bidder_rows.append(row)
            continue

        all_bidders = extract_bidders_list(desc)
        awarded_names = {name.lower() for name in extract_awarded_bidders(desc)}
        for bidder_name in all_bidders:
            won = bidder_name.lower() in awarded_names
            bidder_rows.append({
                "country": notice.get("country"),
                "notice_type": notice.get("notice_type"),
                "status": notice.get("status"),
                "title": notice.get("title"),
                "project_id": notice.get("project_id"),
                "project_name": notice.get("project_name"),
                "procurement_method": notice.get("procurement_method"),
                "borrower": notice.get("borrower"),
                "borrower_bid_reference": notice.get("borrower_bid_reference"),
                "notice_no": notice.get("notice_no"),
                "notice_date": notice.get("notice_date"),
                "submission_date": notice.get("submission_date"),
                "bidder_name": bidder_name,
                "bidder_country": None,
                "bidder_status": "Awarded" if won else "Bidder",
                "won": "Yes" if won else "No",
                "bid_price_at_opening": None,
                "opening_currency": None,
                "evaluated_bid_price": None,
                "evaluated_bid_currency": None,
                "final_evaluation_price": None,
                "final_evaluation_currency": None,
                "winner_contract_amount": notice.get("contract_amount") if won else None,
                "winner_contract_currency": notice.get("currency") if won else None,
                "contact_email": notice.get("contact_email"),
                "url": notice.get("url"),
                "opportunity_id": notice.get("opportunity_id"),
                "description": notice.get("description"),
            })

    return bidder_rows


def _format_template_amount(amount, currency):
    if amount is None:
        return None
    currency_text = (currency or "").strip()
    try:
        number = f"{float(amount):,.2f}"
    except Exception:
        number = str(amount)
    symbols = {
        "USD": "$",
        "US$": "$",
        "EUR": "\u20ac",
        "GBP": "\u00a3",
    }
    prefix = symbols.get(currency_text.upper(), f"{currency_text} " if currency_text else "")
    return f"{prefix}{number}"


def fetch_template_export_rows(where: str, params):
    linked_rows = q(f"""
        WITH filtered_notices AS (
            SELECT *
            FROM procurement_notices
            WHERE {where}
        )
        SELECT
            pn.country,
            b.name AS contractor,
            b.country AS contractor_country,
            pn.title AS tender,
            COALESCE(ba.award_date, pn.notice_date)::text AS award_date,
            COALESCE(pn.submission_date, pn.submission_deadline::date)::text AS deadline_date,
            COALESCE(ba.award_amount, pn.contract_amount) AS award_amount,
            COALESCE(NULLIF(ba.currency, ''), NULLIF(pn.currency, '')) AS award_currency,
            COALESCE(NULLIF(pn.borrower_bid_reference, ''), NULLIF(pn.notice_no, ''), pn.id) AS contractor_reference_number,
            COALESCE(NULLIF(b.contact_email, ''), NULLIF(pn.contact_email, '')) AS contact
        FROM filtered_notices pn
        JOIN bidder_awards ba ON ba.notice_id = pn.id
        JOIN bidders b ON b.id = ba.bidder_id
        WHERE ba.won IS TRUE
        ORDER BY pn.country ASC, COALESCE(ba.award_date, pn.notice_date) DESC NULLS LAST, b.name ASC
    """, params or ())

    if linked_rows:
        return [
            {
                "country": row.get("country"),
                "contractor": row.get("contractor"),
                "contractor_country": row.get("contractor_country"),
                "tender": row.get("tender"),
                "award_date": row.get("award_date"),
                "deadline_date": row.get("deadline_date"),
                "award_amount": _format_template_amount(row.get("award_amount"), row.get("award_currency")),
                "contractor_reference_number": row.get("contractor_reference_number"),
                "contact": row.get("contact"),
            }
            for row in linked_rows
        ]

    rows = fetch_custom_bidder_export_rows(where, params, DEFAULT_CUSTOM_BIDDER_EXPORT_FIELDS)
    output = []
    for row in rows:
        if row.get("won") != "Yes":
            continue
        amount = row.get("winner_contract_amount")
        currency = row.get("winner_contract_currency")
        output.append({
            "country": row.get("country"),
            "contractor": row.get("bidder_name"),
            "contractor_country": row.get("bidder_country"),
            "tender": row.get("title"),
            "award_date": row.get("notice_date"),
            "deadline_date": row.get("submission_date"),
            "award_amount": _format_template_amount(amount, currency),
            "contractor_reference_number": (
                row.get("borrower_bid_reference")
                or row.get("notice_no")
                or row.get("opportunity_id")
            ),
            "contact": row.get("contact_email"),
        })
    return output


def _safe_sheet_title(title: str, fallback: str) -> str:
    cleaned = re.sub(r'[\[\]\:\*\?\/\\]', '', (title or "").strip())
    return (cleaned or fallback)[:31]


def populate_template_sheet(ws, rows, include_deadline: bool = False):
    columns = get_template_export_columns(include_deadline)
    ws.row_dimensions[1].height = 30.6

    for col_idx, (label, key, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=label)
        if label:
            cell.font = Font(color="FFFFFF", size=11, name="Calibri")
            cell.fill = PatternFill("solid", fgColor="356854")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (_, key, _) in enumerate(columns, 1):
            value = row.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=11, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if key in ("award_date", "deadline_date") and value:
                try:
                    if isinstance(value, str) and len(value) >= 10:
                        cell.value = datetime.strptime(value[:10], "%Y-%m-%d").date()
                except Exception:
                    pass
            elif key == "award_amount" and value is not None:
                cell.alignment = Alignment(horizontal="right", vertical="center")


def build_template_workbook(rows, include_deadline: bool = False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    populate_template_sheet(ws, rows, include_deadline)
    return wb


def build_country_template_workbook(rows, countries, include_deadline: bool = False):
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    used_titles = set()
    rows_by_country: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        rows_by_country.setdefault(row.get("country") or "", []).append(row)

    for index, country in enumerate(countries, 1):
        base_title = _safe_sheet_title(country, f"Country {index}")
        title = base_title
        suffix = 2
        while title.lower() in used_titles:
            tail = f" {suffix}"
            title = f"{base_title[:31 - len(tail)]}{tail}"
            suffix += 1
        used_titles.add(title.lower())

        ws = wb.create_sheet(title)
        populate_template_sheet(ws, rows_by_country.get(country, []), include_deadline)

    if not wb.worksheets:
        ws = wb.create_sheet("Sheet1")
        populate_template_sheet(ws, rows, include_deadline)

    return wb


@app.get("/api/export")
def export_excel(
    country: Optional[str] = None,
    notice_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    type: Optional[str] = "normal",
    fields: Optional[str] = None,
    include_deadline: bool = Query(False),
):
    """Export filtered procurement notices to Excel with professional styling and design."""
    where, params = build_where(
        country=country, notice_type=notice_type, status=status,
        search=search, from_date=from_date, to_date=to_date
    )

    is_template_export = type in ("template", "contractor_template")
    is_country_template_export = type in ("template_by_country", "contractor_template_by_country")
    is_huzalink_export = type == "huzalink"
    is_custom_bidder_export = bool(fields) and not is_huzalink_export

    if is_template_export or is_country_template_export:
        rows = fetch_template_export_rows(where, params)
        if is_country_template_export:
            configured_rows = q("SELECT name FROM target_countries ORDER BY name")
            countries = [row["name"] for row in configured_rows] or DEFAULT_COUNTRIES
            wb = build_country_template_workbook(rows, countries, include_deadline)
        else:
            wb = build_template_workbook(rows, include_deadline)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
        deadline_part = "_With_Deadline" if include_deadline else ""
        filename = (
            f"Contractor_Template_By_Country{deadline_part}_{timestamp}.xlsx"
            if is_country_template_export else
            f"Contractor_Template_Export{deadline_part}_{timestamp}.xlsx"
        )
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    elif is_huzalink_export:
        selected_fields = resolve_export_fields(fields)
        rows = fetch_export_rows(where, params, selected_fields)
    elif is_custom_bidder_export:
        selected_fields = resolve_custom_bidder_export_fields(fields)
        rows = fetch_custom_bidder_export_rows(where, params, selected_fields)
    else:
        selected_fields = DEFAULT_EXPORT_FIELDS.copy()
        rows = fetch_export_rows(where, params, DEFAULT_EXPORT_FIELDS + ['description'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procurement Notices"

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    header_border = Border(
        left=Side(style="thin", color="34495E"),
        right=Side(style="thin", color="34495E"),
        top=Side(style="medium", color="3498DB"),
        bottom=Side(style="medium", color="3498DB")
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(size=11, name="Calibri")
    border_all = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7")
    )
    link_font = Font(color="3498DB", underline="single", size=11, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if is_huzalink_export:
        COLUMNS = [EXPORT_FIELD_CONFIG[field] for field in selected_fields if field in EXPORT_FIELD_CONFIG]
        ws.title = "Huzalink Export - Custom Fields"
    elif is_custom_bidder_export:
        COLUMNS = [CUSTOM_BIDDER_EXPORT_FIELD_CONFIG[field] for field in selected_fields if field in CUSTOM_BIDDER_EXPORT_FIELD_CONFIG]
        ws.title = "Custom Bidder Export"
    else:
        COLUMNS = (
            [EXPORT_FIELD_CONFIG[field] for field in DEFAULT_EXPORT_FIELDS]
            + [EXPORT_FIELD_CONFIG["overview"], EXPORT_FIELD_CONFIG["requirements"], ("Description", "description", 80)]
        )
        ws.title = "Procurement Notices"

    data_start_row = 1
    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.sheet_view.zoomScale = 85
    data_start_row = 2

    for row_idx, row in enumerate(rows, data_start_row):
        ws.row_dimensions[row_idx].height = 25
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            if key == 'bidders':
                desc = row.get('description')
                val = extract_bidders_from_description(desc) if desc else None
            elif key == "overview":
                val = extract_notice_overview(row.get("description"))
            elif key == "requirements":
                val = extract_notice_requirements(row.get("description"))
            elif key == "is_tech":
                val = "Yes" if classify_notice_tech(row).get("is_tech") else "No"
            elif key == "tech_category":
                val = classify_notice_tech(row).get("tech_category")
            else:
                val = row.get(key)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border_all
            cell.alignment = left_align

            if key == "url" and val:
                cell.hyperlink = val
                cell.font = link_font
                cell.value = "View on World Bank"
                cell.alignment = center_align
            elif key in ("description", "overview", "requirements") and val:
                desc_text = str(val)
                if len(desc_text) > 500:
                    desc_text = desc_text[:500] + "..."
                cell.value = desc_text
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif key in ("notice_date", "submission_date", "awarded_date"):
                cell.alignment = center_align
                if val:
                    try:
                        if isinstance(val, str) and len(val) >= 10:
                            cell.value = datetime.strptime(val[:10], "%Y-%m-%d").date()
                    except Exception:
                        pass
            elif key in ("currency", "notice_type", "status", "won", "bidder_status", "is_tech", "opening_currency", "evaluated_bid_currency", "final_evaluation_currency", "winner_contract_currency"):
                cell.alignment = center_align
                if key == "status" and val:
                    if val == "Active":
                        cell.fill = PatternFill("solid", fgColor="E8F5E8")
                    elif val == "Awarded":
                        cell.fill = PatternFill("solid", fgColor="E3F2FD")
                    elif val == "Closed":
                        cell.fill = PatternFill("solid", fgColor="FFF3E0")
                    elif val == "Cancelled":
                        cell.fill = PatternFill("solid", fgColor="FFEBEE")
            elif key in ("contract_amount", "bid_price_at_opening", "evaluated_bid_price", "final_evaluation_price", "winner_contract_amount") and val is not None:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    amount = float(val)
                    if amount >= 1000000:
                        cell.number_format = '$#,##0.0,, "M"'
                    elif amount >= 1000:
                        cell.number_format = '$#,##0.0, "K"'
                    else:
                        cell.number_format = '$#,##0.00'
                except Exception:
                    cell.number_format = '#,##0.00'
            elif key == "contact_email" and val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(color="3498DB", underline="single", size=11, name="Calibri")

    ws2 = wb.create_sheet("Summary & Filters")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 20

    title_cell = ws2.cell(row=1, column=1, value="World Bank Procurement Export Summary")
    title_cell.font = Font(bold=True, size=16, color="2C3E50", name="Calibri")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells("A1:C1")

    export_type_name = (
        "Huzalink Export"
        if is_huzalink_export else
        ("Custom Bidder Export" if is_custom_bidder_export else "Normal Export")
    )
    ws2.cell(row=3, column=1, value=f"{export_type_name} Information").font = Font(bold=True, size=14, color="2C3E50", name="Calibri")
    ws2.cell(row=3, column=1).fill = PatternFill("solid", fgColor="ECF0F1")
    ws2.merge_cells("A3:C3")

    info_data = [
        ("Export Type:", export_type_name),
        ("Export Date & Time:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Total Rows:", len(rows)),
        ("Country Filter:", country or "All Countries"),
        ("Notice Type Filter:", notice_type or "All Types"),
        ("Status Filter:", status or "All Statuses"),
        ("Date Range From:", str(from_date) if from_date else "Not specified"),
        ("Date Range To:", str(to_date) if to_date else "Not specified"),
        ("Search Term:", search or "None"),
    ]
    if is_huzalink_export or is_custom_bidder_export:
        info_data.extend([
            ("Selected Fields:", ", ".join(selected_fields)),
            ("Print Ready:", "Optimized for sharing"),
        ])

    for r, (label, value) in enumerate(info_data, 4):
        label_cell = ws2.cell(row=r, column=1, value=label)
        label_cell.font = Font(bold=True, size=11, color="34495E", name="Calibri")
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        value_cell = ws2.cell(row=r, column=2, value=str(value))
        value_cell.font = Font(size=11, color="2C3E50", name="Calibri")
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws2.cell(row=len(info_data) + 6, column=1, value="Quick Statistics").font = Font(bold=True, size=14, color="2C3E50", name="Calibri")
    ws2.cell(row=len(info_data) + 6, column=1).fill = PatternFill("solid", fgColor="ECF0F1")
    ws2.merge_cells(f"A{len(info_data) + 6}:C{len(info_data) + 6}")

    active_count = sum(1 for r in rows if r.get('status') == 'Active')
    awarded_count = sum(1 for r in rows if r.get('status') == 'Awarded')
    stats_data = [
        ("Active Opportunities:", active_count),
        ("Awarded Opportunities:", awarded_count),
        ("Countries Represented:", len(set(r.get('country') for r in rows if r.get('country')))),
        ("Average Contract Amount:", "N/A"),
    ]
    for r, (label, value) in enumerate(stats_data, len(info_data) + 7):
        label_cell = ws2.cell(row=r, column=1, value=label)
        label_cell.font = Font(bold=True, size=11, color="34495E", name="Calibri")
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        value_cell = ws2.cell(row=r, column=2, value=str(value))
        value_cell.font = Font(size=11, color="2C3E50", name="Calibri")
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    if is_huzalink_export:
        parts = ["Huzalink_Procurement", timestamp]
    elif is_custom_bidder_export:
        parts = ["WB_Procurement_Bidders", timestamp]
    else:
        parts = ["WB_Procurement", timestamp]
    if country: parts.append(country.replace(" ", "_")[:20])
    if notice_type: parts.append(notice_type[:15])
    if status: parts.append(status[:15])
    filename = "_".join(parts) + ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/export/csv")
def export_csv(
    country: Optional[str] = None,
    notice_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    fields: Optional[str] = None,
):
    where, params = build_where(
        country=country, notice_type=notice_type, status=status,
        search=search, from_date=from_date, to_date=to_date
    )
    is_custom_bidder_export = bool(fields)
    if is_custom_bidder_export:
        selected_fields = resolve_custom_bidder_export_fields(fields)
        rows = fetch_custom_bidder_export_rows(where, params, selected_fields)
    else:
        selected_fields = resolve_export_fields(fields)
        rows = fetch_export_rows(where, params, selected_fields)

    output = io.StringIO()
    writer = csv.writer(output)
    field_config = CUSTOM_BIDDER_EXPORT_FIELD_CONFIG if is_custom_bidder_export else EXPORT_FIELD_CONFIG
    headers = [field_config[field][0] for field in selected_fields if field in field_config]
    writer.writerow(headers)
    for row in rows:
        values = []
        for field in selected_fields:
            if not is_custom_bidder_export and field == "bidders":
                values.append(extract_bidders_from_description(row.get("description")) if row.get("description") else None)
            elif field == "overview":
                values.append(extract_notice_overview(row.get("description")))
            elif field == "requirements":
                values.append(extract_notice_requirements(row.get("description")))
            elif field == "is_tech":
                values.append("Yes" if classify_notice_tech(row).get("is_tech") else "No")
            elif field == "tech_category":
                values.append(classify_notice_tech(row).get("tech_category"))
            else:
                values.append(row.get(field_config[field][1]))
        writer.writerow(values)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"{'WB_Procurement_Bidders' if is_custom_bidder_export else 'WB_Procurement'}_{timestamp}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ── Migration ─────────────────────────────────────────────────────────────────

@app.post("/api/migrate")
def migrate_db():
    """Bootstrap and migrate the DB without dropping data. Safe to re-run."""
    try:
        import fetcher
        fetcher.init_db()
    except Exception as e:
        raise HTTPException(500, f"Base schema initialization failed: {e}")

    ensure_support_tables()

    migrations = [
        "ALTER TABLE target_countries ADD COLUMN IF NOT EXISTS added_at TIMESTAMPTZ DEFAULT NOW()",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS notice_no TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS notice_status TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS language TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS borrower_bid_reference TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS submission_deadline TIMESTAMPTZ",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contact_name TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contact_org TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contact_address TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contact_city TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contact_phone TEXT",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contact_website TEXT",
        "UPDATE procurement_notices SET notice_status = status WHERE notice_status IS NULL",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS contract_amount NUMERIC",
        "ALTER TABLE procurement_notices ADD COLUMN IF NOT EXISTS currency TEXT",
        "ALTER TABLE bidders ADD COLUMN IF NOT EXISTS linkedin_url TEXT",
        "ALTER TABLE bidders ADD COLUMN IF NOT EXISTS business_model TEXT",
        "ALTER TABLE bidders ADD COLUMN IF NOT EXISTS core_products TEXT",
        "ALTER TABLE bidders ADD COLUMN IF NOT EXISTS corporate_activities TEXT",
        """CREATE TABLE IF NOT EXISTS bidders (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            category TEXT,
            contact_name TEXT,
            contact_email TEXT,
            contact_phone TEXT,
            linkedin_url TEXT,
            contact_org TEXT,
            country TEXT,
            business_model TEXT,
            core_products TEXT,
            corporate_activities TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS bidder_awards (
            id SERIAL PRIMARY KEY,
            bidder_id INTEGER NOT NULL REFERENCES bidders(id) ON DELETE CASCADE,
            notice_id TEXT NOT NULL REFERENCES procurement_notices(id) ON DELETE CASCADE,
            won BOOLEAN DEFAULT FALSE,
            award_amount NUMERIC,
            currency TEXT,
            award_date DATE,
            role TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(bidder_id, notice_id)
        )""",
    ]
    with db() as conn:
        with conn.cursor() as cur:
            for stmt in migrations:
                cur.execute(stmt)
        conn.commit()
    return {
        "status": "ok",
        "message": "Database bootstrapped and migrations applied.",
        "migrations": len(migrations),
    }


# ── Bidder endpoints ──────────────────────────────────────────────────────────

@app.post("/api/bidders/from_notice")
def import_bidders_from_notice(notice_id: str = Query(None), fetch_detail: bool = Query(False)):
    """
    Extract bidders from a stored notice and upsert into `bidders` and `bidder_awards`.

    IMPORTANT: The `borrower` field is the HOST INSTITUTION (government client, e.g.
    "Ministry of Finance"). It is NEVER a bidder. Only companies found in the notice
    description text or scraped from the WB detail page are stored as bidders.
    """
    if not notice_id:
        raise HTTPException(status_code=400, detail="notice_id is required")

    rows = q("SELECT * FROM procurement_notices WHERE id = %s", [notice_id])
    if not rows:
        raise HTTPException(status_code=404, detail="Notice not found")

    notice = rows[0]
    desc = notice.get('description') or ''
    bidder_details = parse_notice_bidder_details(desc)
    details_by_name = {
        detail["name"].lower(): detail
        for detail in bidder_details
        if detail.get("name")
    }
    inferred_category = _infer_notice_category(notice, desc)

    # Extract all participating bidders and explicitly awarded winners
    all_bidders = [detail["name"] for detail in bidder_details] or extract_bidders_list(desc)
    awarded = [
        detail["name"] for detail in bidder_details
        if detail.get("section") == "awarded"
    ] or extract_awarded_bidders(desc)

    # Optionally scrape the WB detail page for richer bidder data
    if fetch_detail:
        detail_bidders = fetch_bidders_from_detail_page(notice_id, notice.get('url') or '')
        for b in detail_bidders:
            if b.lower() not in {x.lower() for x in all_bidders}:
                all_bidders.append(b)

    # If nothing found — return cleanly. Do NOT fall back to borrower.
    # The borrower is the CLIENT, not a bidding company.
    if not all_bidders:
        return {
            "status": "ok",
            "bidders_found": 0,
            "bidders_inserted": 0,
            "links_created": 0,
            "note": (
                "No bidder names found in the description or detail page. "
                "This notice may not contain structured bidder data. "
                "Try 'Fetch from World Bank Page' for richer data."
            )
        }

    inserted = 0
    linked = 0

    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT bidder_id FROM bidder_awards WHERE notice_id = %s",
                (notice_id,)
            )
            previous_bidder_ids = [
                row["bidder_id"] if hasattr(row, "__getitem__") else row[0]
                for row in cur.fetchall()
            ]
            cur.execute("DELETE FROM bidder_awards WHERE notice_id = %s", (notice_id,))

            for name in all_bidders:
                detail = details_by_name.get(name.lower(), {})
                bidder_country = detail.get("country")
                bidder_category = inferred_category
                bid_amount = detail.get("amount")
                bid_currency = detail.get("currency") or notice.get('currency')

                # Upsert bidder
                cur.execute(
                    "SELECT id, category, country FROM bidders WHERE lower(name) = lower(%s)",
                    (name,)
                )
                existing = cur.fetchone()
                if existing:
                    bidder_id = existing["id"] if hasattr(existing, "__getitem__") else existing[0]
                    existing_category = existing.get("category") if hasattr(existing, "get") else None
                    existing_country = existing.get("country") if hasattr(existing, "get") else None
                    cur.execute(
                        """UPDATE bidders
                           SET category = %s,
                               country = COALESCE(NULLIF(country, ''), %s),
                               updated_at = NOW()
                           WHERE id = %s""",
                        (_merge_categories(existing_category, bidder_category), bidder_country, bidder_id)
                    )
                else:
                    cur.execute(
                        """INSERT INTO bidders (name, category, country, created_at, updated_at)
                           VALUES (%s, %s, %s, NOW(), NOW())
                           RETURNING id""",
                        (name, bidder_category, bidder_country)
                    )
                    row = cur.fetchone()
                    bidder_id = row["id"] if hasattr(row, "__getitem__") else row[0]
                    inserted += 1

                # won=True only when explicitly named as awarded in the text
                # Never derived from the borrower field
                won = detail.get("section") == "awarded" or any(name.lower() == w.lower() for w in awarded)

                # Link to notice
                cur.execute(
                    "SELECT id FROM bidder_awards WHERE bidder_id = %s AND notice_id = %s",
                    (bidder_id, notice_id)
                )
                link = cur.fetchone()
                if link:
                    link_id = link["id"] if hasattr(link, "__getitem__") else link[0]
                    cur.execute(
                        """UPDATE bidder_awards
                           SET won = %s, award_amount = %s, currency = %s,
                               award_date = %s, role = %s, updated_at = NOW()
                           WHERE id = %s""",
                        (won, bid_amount or notice.get('contract_amount'), bid_currency,
                         notice.get('notice_date'), detail.get("role"), link_id)
                    )
                else:
                    cur.execute(
                        """INSERT INTO bidder_awards
                           (bidder_id, notice_id, won, award_amount, currency,
                            award_date, role, created_at, updated_at)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())""",
                        (bidder_id, notice_id, won,
                         bid_amount or notice.get('contract_amount'), bid_currency,
                         notice.get('notice_date'), detail.get("role"))
                    )
                    linked += 1

            if previous_bidder_ids:
                cur.execute(
                    """DELETE FROM bidders b
                       WHERE b.id = ANY(%s)
                         AND NOT EXISTS (
                             SELECT 1 FROM bidder_awards ba
                             WHERE ba.bidder_id = b.id
                         )""",
                    (previous_bidder_ids,)
                )

        conn.commit()

    return {
        "status": "ok",
        "bidders_found": len(all_bidders),
        "bidders_inserted": inserted,
        "links_created": linked,
        "awarded_identified": len(awarded),
    }


@app.get("/api/bidders")
def list_bidders(
    qs:       Optional[str] = Query(None,  description="Search by name (legacy param)"),
    search:   Optional[str] = Query(None,  description="Search by name or org"),
    country:  Optional[str] = Query(None,  description="Filter by country"),
    won_only: bool          = Query(False, description="Only bidders who won at least one award"),
    page:     int           = Query(1,  ge=1),
    page_size:int           = Query(25, ge=1, le=200),
):
    """List bidders with aggregated bid stats."""
    where, params = build_bidder_filters(search, qs, country)
    having = "HAVING COUNT(CASE WHEN ba.won THEN 1 END) > 0" if won_only else ""
    offset = (page - 1) * page_size

    rows = q(f"""
        SELECT
            b.id,
            b.name,
            b.category,
            b.contact_name,
            b.contact_email,
            b.contact_phone,
            b.contact_org,
            b.country,
            b.business_model,
            b.core_products,
            b.corporate_activities,
            b.created_at,
            b.updated_at,
            COUNT(ba.id)                              AS bid_count,
            COUNT(CASE WHEN ba.won THEN 1 END)        AS won_count,
            COALESCE(SUM(ba.award_amount), 0)         AS total_bid_amount,
            MAX(NULLIF(ba.currency, ''))              AS primary_currency,
            MAX(ba.award_date)::text                  AS last_bid_date,
            (SELECT pn.title
             FROM bidder_awards ba2
             JOIN procurement_notices pn ON pn.id = ba2.notice_id
             WHERE ba2.bidder_id = b.id
             ORDER BY ba2.award_date DESC NULLS LAST
             LIMIT 1)                                 AS latest_bid_title
        FROM bidders b
        LEFT JOIN bidder_awards ba ON ba.bidder_id = b.id
        WHERE {where}
        GROUP BY b.id
        {having}
        ORDER BY bid_count DESC, b.name ASC
        LIMIT %s OFFSET %s
    """, params + [page_size, offset])

    for row in rows:
        row["is_tech"] = looks_like_tech_bidder(row)

    total_rows = q(f"""
        SELECT COUNT(*) AS cnt FROM (
            SELECT b.id
            FROM bidders b
            LEFT JOIN bidder_awards ba ON ba.bidder_id = b.id
            WHERE {where}
            GROUP BY b.id
            {having}
        ) sub
    """, params)
    total = total_rows[0]["cnt"] if total_rows else 0

    return {
        "page":        page,
        "page_size":   page_size,
        "total":       total,
        "total_pages": max(1, -(-total // page_size)),
        "data":        [dict(r) for r in rows],
    }


BIDDER_EXPORT_FIELDS = {
    "name": ("Bidder Name", "name", 40),
    "country": ("Country of Origin", "country", 20),
    "category": ("Category", "category", 24),
    "bid_count": ("Total Bids", "bid_count", 14),
    "won_count": ("Won Bids", "won_count", 14),
    "total_bid_amount": ("Total Bid Amount", "total_bid_amount", 18),
    "primary_currency": ("Currency", "primary_currency", 12),
    "last_bid_date": ("Last Bid Date", "last_bid_date", 16),
    "latest_bid_title": ("Latest Bid", "latest_bid_title", 50),
    "contact_name": ("Contact Name", "contact_name", 24),
    "contact_email": ("Contact Email", "contact_email", 30),
    "contact_phone": ("Contact Phone", "contact_phone", 20),
    "contact_org": ("Organisation", "contact_org", 30),
    "business_model": ("Business Model", "business_model", 30),
    "core_products": ("Core Products", "core_products", 30),
    "corporate_activities": ("Corporate Activities", "corporate_activities", 40),
}


def resolve_bidder_export_fields(fields: Optional[str]) -> List[str]:
    selected_fields = [f.strip() for f in (fields or "").split(",") if f.strip()]
    if not selected_fields:
        return ["name", "country", "category", "bid_count", "won_count", "total_bid_amount", "primary_currency", "last_bid_date"]
    return [field for field in selected_fields if field in BIDDER_EXPORT_FIELDS]


def fetch_bidder_export_rows(search: Optional[str], qs: Optional[str], country: Optional[str], won_only: bool):
    where, params = build_bidder_filters(search, qs, country)
    having = "HAVING COUNT(CASE WHEN ba.won THEN 1 END) > 0" if won_only else ""
    return q(f"""
        SELECT
            b.name,
            b.country,
            b.category,
            b.contact_name,
            b.contact_email,
            b.contact_phone,
            b.contact_org,
            b.business_model,
            b.core_products,
            b.corporate_activities,
            COUNT(ba.id)                       AS bid_count,
            COUNT(CASE WHEN ba.won THEN 1 END) AS won_count,
            COALESCE(SUM(ba.award_amount), 0)  AS total_bid_amount,
            MAX(NULLIF(ba.currency, ''))       AS primary_currency,
            MAX(ba.award_date)::text           AS last_bid_date,
            (SELECT pn.title
             FROM bidder_awards ba2
             JOIN procurement_notices pn ON pn.id = ba2.notice_id
             WHERE ba2.bidder_id = b.id
             ORDER BY ba2.award_date DESC NULLS LAST
             LIMIT 1) AS latest_bid_title
        FROM bidders b
        LEFT JOIN bidder_awards ba ON ba.bidder_id = b.id
        WHERE {where}
        GROUP BY b.id
        {having}
        ORDER BY bid_count DESC, b.name ASC
    """, params)


@app.get("/api/bidders/export")
def export_bidders_excel(
    qs: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    won_only: bool = Query(False),
    fields: Optional[str] = Query(None),
):
    selected_fields = resolve_bidder_export_fields(fields)
    rows = fetch_bidder_export_rows(search, qs, country, won_only)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bidders"

    for col_idx, field in enumerate(selected_fields, 1):
        label, _, width = BIDDER_EXPORT_FIELDS[field]
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5F43")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(selected_fields, 1):
            value = row.get(BIDDER_EXPORT_FIELDS[field][1])
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Bidders_Custom_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/bidders/export/csv")
def export_bidders_csv(
    qs: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    won_only: bool = Query(False),
    fields: Optional[str] = Query(None),
):
    selected_fields = resolve_bidder_export_fields(fields)
    rows = fetch_bidder_export_rows(search, qs, country, won_only)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([BIDDER_EXPORT_FIELDS[field][0] for field in selected_fields])
    for row in rows:
        writer.writerow([row.get(BIDDER_EXPORT_FIELDS[field][1]) for field in selected_fields])

    filename = f"Bidders_Custom_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/bidders/import_status")
def get_bidder_import_status():
    total_awards = q("SELECT COUNT(*) AS cnt FROM procurement_notices WHERE notice_type IN ('Contract Award','Award') OR status = 'Awarded'")
    linked_awards = q("SELECT COUNT(DISTINCT notice_id) AS cnt FROM bidder_awards")
    missing_awards = q("""
        SELECT COUNT(*) AS cnt
        FROM procurement_notices pn
        WHERE (pn.notice_type IN ('Contract Award','Award') OR pn.status = 'Awarded')
          AND NOT EXISTS (SELECT 1 FROM bidder_awards ba WHERE ba.notice_id = pn.id)
    """)
    sample_missing = q("""
        SELECT pn.id, pn.title, pn.notice_date::text
        FROM procurement_notices pn
        WHERE (pn.notice_type IN ('Contract Award','Award') OR pn.status = 'Awarded')
          AND NOT EXISTS (SELECT 1 FROM bidder_awards ba WHERE ba.notice_id = pn.id)
        ORDER BY pn.notice_date DESC NULLS LAST
        LIMIT 10
    """)

    summary = None
    try:
        with open('bidders_import_summary.json', 'r', encoding='utf-8') as f:
            summary = json.loads(f.read())
    except Exception:
        summary = None

    return {
        "total_award_notices": total_awards[0]["cnt"] if total_awards else 0,
        "linked_award_notices": linked_awards[0]["cnt"] if linked_awards else 0,
        "missing_award_notices": missing_awards[0]["cnt"] if missing_awards else 0,
        "missing_samples": [dict(r) for r in sample_missing],
        "last_summary": summary,
    }


@app.post("/api/bidders/import_missing")
def import_missing_awards(fetch_detail: bool = Query(False)):
    notices = q("""
        SELECT pn.id
        FROM procurement_notices pn
        WHERE (pn.notice_type IN ('Contract Award','Award') OR pn.status = 'Awarded')
          AND NOT EXISTS (SELECT 1 FROM bidder_awards ba WHERE ba.notice_id = pn.id)
        ORDER BY pn.notice_date DESC NULLS LAST
    """)
    summary = {"processed": 0, "bidders_found": 0, "bidders_inserted": 0, "links_created": 0, "errors": 0}
    for row in notices:
        try:
            res = import_bidders_from_notice(row["id"], fetch_detail)
            summary["processed"] += 1
            summary["bidders_found"] += res.get("bidders_found", 0)
            summary["bidders_inserted"] += res.get("bidders_inserted", 0)
            summary["links_created"] += res.get("links_created", 0)
        except Exception:
            summary["errors"] += 1
    return summary


@app.get("/api/bidders/countries")
def list_bidder_countries():
    rows = q("""
        SELECT DISTINCT country
        FROM bidders
        WHERE country IS NOT NULL AND TRIM(country) <> ''
        ORDER BY country
    """)
    return [row["country"] for row in rows]


@app.get("/api/bidders/{bidder_id}")
def get_bidder(bidder_id: int):
    rows = q("SELECT * FROM bidders WHERE id = %s", [bidder_id])
    if not rows:
        raise HTTPException(status_code=404, detail="Bidder not found")
    return dict(rows[0])


@app.get("/api/bidders/{bidder_id}/notices")
def get_bidder_notices(bidder_id: int):
    """All procurement notices a bidder participated in, with bid amounts and win status."""
    rows = q("""
        SELECT
            pn.id,
            pn.title,
            pn.project_name,
            pn.project_id,
            pn.country          AS borrower_country,
            pn.notice_type,
            pn.notice_date::text,
            COALESCE(pn.submission_date, pn.submission_deadline::date)::text AS submission_date,
            pn.status,
            pn.url,
            pn.contract_amount  AS notice_contract_amount,
            pn.currency         AS notice_currency,
            pn.borrower,
            ba.won,
            ba.award_amount,
            ba.currency,
            COALESCE(ba.award_amount, pn.contract_amount) AS bid_amount,
            COALESCE(NULLIF(ba.currency, ''), NULLIF(pn.currency, '')) AS bid_currency,
            ba.award_date::text,
            ba.role
        FROM bidder_awards ba
        JOIN procurement_notices pn ON pn.id = ba.notice_id
        WHERE ba.bidder_id = %s
        ORDER BY ba.award_date DESC NULLS LAST, pn.notice_date DESC NULLS LAST
    """, [bidder_id])
    return [dict(r) for r in rows]


class BidderUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    linkedin_url: Optional[str] = None
    contact_org: Optional[str] = None
    country: Optional[str] = None
    business_model: Optional[str] = None
    core_products: Optional[str] = None
    corporate_activities: Optional[str] = None


@app.put("/api/bidders/{bidder_id}")
def update_bidder(bidder_id: int, body: BidderUpdate):
    fields = {k: v for k, v in body.dict().items() if v is not None}
    if not fields:
        return {"status": "ok"}
    sets = []
    params = []
    for k, v in fields.items():
        sets.append(f"{k} = %s")
        params.append(v)
    params.append(bidder_id)
    sql = f"UPDATE bidders SET {', '.join(sets)}, updated_at = NOW() WHERE id = %s"
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
        conn.commit()
    return {"status": "ok"}


@app.delete("/api/bidders/{bidder_id}")
def delete_bidder(bidder_id: int):
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM bidders WHERE id = %s", [bidder_id])
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Bidder not found")
            cur.execute("DELETE FROM bidders WHERE id = %s", [bidder_id])
        conn.commit()
    return {"status": "ok", "deleted_id": bidder_id, "deleted_name": row["name"]}


def build_bidder_filters(search: Optional[str], qs: Optional[str], country: Optional[str]):
    search_term = search or qs
    filters = ["1=1"]
    params: List[Any] = []
    if search_term:
        filters.append("(b.name ILIKE %s OR b.contact_org ILIKE %s OR b.contact_name ILIKE %s)")
        like = f"%{search_term}%"
        params.extend([like, like, like])
    if country:
        filters.append("b.country = %s")
        params.append(country)
    return " AND ".join(filters), params


@app.get("/api/notices/{notice_id}/bidders")
def get_notice_bidders(notice_id: str):
    """All bidders linked to a specific notice."""
    rows = q(
        """SELECT b.*, ba.won, ba.award_amount, ba.currency, ba.award_date
           FROM bidder_awards ba
           JOIN bidders b ON b.id = ba.bidder_id
           WHERE ba.notice_id = %s
           ORDER BY ba.won DESC, b.name ASC""",
        [notice_id]
    )
    return [dict(r) for r in rows]


@app.get("/api/notices/awards")
def get_award_notices(
    country: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, le=100),
):
    """
    List Contract Award notices with their bidder counts.
    Useful for seeing which notices have been processed and how many bidders were found.
    """
    filters = ["(notice_type = 'Contract Award' OR notice_type = 'Award')"]
    params = []

    if country:
        filters.append("pn.country = %s")
        params.append(country)
    if search:
        filters.append("(pn.title ILIKE %s OR pn.project_name ILIKE %s)")
        like = f"%{search}%"
        params.extend([like, like])

    where = " AND ".join(filters)
    offset = (page - 1) * page_size

    rows = q(f"""
        SELECT
            pn.id,
            pn.country,
            pn.title,
            pn.project_id,
            pn.project_name,
            pn.notice_date::text,
            pn.contract_amount,
            pn.currency,
            pn.borrower,
            pn.status,
            pn.url,
            COUNT(ba.id) AS bidder_count,
            COUNT(CASE WHEN ba.won THEN 1 END) AS winner_count
        FROM procurement_notices pn
        LEFT JOIN bidder_awards ba ON ba.notice_id = pn.id
        WHERE {where}
        GROUP BY pn.id
        ORDER BY pn.notice_date DESC NULLS LAST
        LIMIT %s OFFSET %s
    """, params + [page_size, offset])

    total = q(
        f"SELECT COUNT(*) AS cnt FROM procurement_notices pn WHERE {where}",
        params
    )

    return {
        "total": total[0]["cnt"] if total else 0,
        "page": page,
        "page_size": page_size,
        "data": [dict(r) for r in rows],
    }


@app.post("/api/bidders/import_awards_all")
def import_awards_all(
    fetch_detail: bool = Query(False),
    country: Optional[str] = Query(None),
):
    """Batch import bidders for all award notices, optionally limited to one country. Runs in background."""
    def run_batch():
        if country:
            notices = q(
                """SELECT id
                   FROM procurement_notices
                   WHERE (notice_type IN ('Contract Award','Award') OR status = 'Awarded')
                     AND country = %s""",
                [country]
            )
        else:
            notices = q("SELECT id FROM procurement_notices WHERE notice_type IN ('Contract Award','Award') OR status = 'Awarded'")
        summary = {
            "processed": 0,
            "bidders_found": 0,
            "bidders_inserted": 0,
            "links_created": 0,
            "country": country,
        }
        for n in notices:
            nid = n['id']
            try:
                res = import_bidders_from_notice(nid, fetch_detail)
                summary['processed'] += 1
                summary['bidders_found'] += res.get('bidders_found', 0)
                summary['bidders_inserted'] += res.get('bidders_inserted', 0)
                summary['links_created'] += res.get('links_created', 0)
            except Exception:
                continue
        try:
            with open('bidders_import_summary.json', 'w', encoding='utf-8') as f:
                f.write(json.dumps(summary))
        except Exception:
            pass

    threading.Thread(target=run_batch, daemon=True).start()
    return {"status": "started", "country": country}


@app.post("/api/bidders/import_by_country")
def import_bidders_by_country(
    country: str = Query(..., description="Import bidders only for this country"),
    fetch_detail: bool = Query(False),
):
    if not country or not country.strip():
        raise HTTPException(status_code=400, detail="country is required")
    return import_awards_all(fetch_detail=fetch_detail, country=country.strip())


@app.post("/api/bidders/import_missing_by_country")
def import_missing_awards_by_country(
    country: str = Query(..., description="Import missing bidder links only for this country"),
    fetch_detail: bool = Query(False),
):
    normalized_country = (country or "").strip()
    if not normalized_country:
        raise HTTPException(status_code=400, detail="country is required")

    notices = q("""
        SELECT pn.id
        FROM procurement_notices pn
        WHERE (pn.notice_type IN ('Contract Award','Award') OR pn.status = 'Awarded')
          AND pn.country = %s
          AND NOT EXISTS (SELECT 1 FROM bidder_awards ba WHERE ba.notice_id = pn.id)
        ORDER BY pn.notice_date DESC NULLS LAST
    """, [normalized_country])
    summary = {
        "processed": 0,
        "bidders_found": 0,
        "bidders_inserted": 0,
        "links_created": 0,
        "errors": 0,
        "country": normalized_country,
    }
    for row in notices:
        try:
            res = import_bidders_from_notice(row["id"], fetch_detail)
            summary["processed"] += 1
            summary["bidders_found"] += res.get("bidders_found", 0)
            summary["bidders_inserted"] += res.get("bidders_inserted", 0)
            summary["links_created"] += res.get("links_created", 0)
        except Exception:
            summary["errors"] += 1
    return summary


# ── Contact Enrichment ─────────────────────────────────────────────────────────

class EnrichResult(BaseModel):
    status: str
    bidder_id: int
    bidder_name: str
    found: Dict[str, Any] = {}
    updated_fields: List[str] = []
    error: Optional[str] = None


@app.post("/api/bidders/{bidder_id}/enrich")
def enrich_bidder_contact(bidder_id: int):
    rows = q("SELECT * FROM bidders WHERE id = %s", [bidder_id])
    if not rows:
        raise HTTPException(status_code=404, detail="Bidder not found")
    bidder = rows[0]
    company = bidder["name"]
    country = bidder.get("country") or ""

    try:
        found = search_company_contact(company, country)
    except Exception as e:
        return EnrichResult(
            status="error", bidder_id=bidder_id, bidder_name=company,
            error=str(e)
        ).dict()

    if not found:
        return EnrichResult(
            status="no_results", bidder_id=bidder_id, bidder_name=company,
            found={}, updated_fields=[]
        ).dict()

    update_fields = {}
    for key in ("contact_email", "contact_phone", "linkedin_url"):
        val = found.get(key)
        if val:
            update_fields[key] = val

    if update_fields:
        sets = ", ".join(f"{k} = %s" for k in update_fields)
        params = list(update_fields.values()) + [bidder_id]
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE bidders SET {sets}, updated_at = NOW() WHERE id = %s", params)
            conn.commit()

    return EnrichResult(
        status="ok", bidder_id=bidder_id, bidder_name=company,
        found=found, updated_fields=list(update_fields.keys())
    ).dict()


@app.post("/api/bidders/{bidder_id}/enrich_gemini")
def enrich_bidder_gemini(bidder_id: int):
    # 1. Fetch bidder details
    rows = q("SELECT * FROM bidders WHERE id = %s", [bidder_id])
    if not rows:
        raise HTTPException(status_code=404, detail="Bidder not found")
    bidder = rows[0]
    
    # 2. Fetch bidder notices for context
    notices = get_bidder_notices(bidder_id)
    
    # 3. Call Gemini service
    try:
        enriched = classify_and_enrich_with_gemini(
            company=bidder["name"],
            country=bidder.get("country") or "",
            bids=notices
        )
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
    # 4. Save results to DB
    update_fields = {}
    for key in ("category", "business_model", "core_products", "corporate_activities", "contact_email", "contact_phone", "linkedin_url"):
        val = enriched.get(key)
        if val is not None:
            update_fields[key] = val
            
    if update_fields:
        sets = ", ".join(f"{k} = %s" for k in update_fields)
        params = list(update_fields.values()) + [bidder_id]
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE bidders SET {sets}, updated_at = NOW() WHERE id = %s", params)
            conn.commit()
            
    return {
        "status": "ok",
        "bidder_id": bidder_id,
        "bidder_name": bidder["name"],
        "enriched": enriched,
        "updated_fields": list(update_fields.keys())
    }


@app.post("/api/bidders/enrich")
def enrich_all_bidders(
    missing_only: bool = Query(True, description="Only enrich bidders missing contact info"),
    limit: int = Query(50, ge=1, le=500),
):
    if missing_only:
        rows = q("""
            SELECT id, name, country FROM bidders
            WHERE (contact_email IS NULL OR TRIM(contact_email) = '')
              AND (contact_phone IS NULL OR TRIM(contact_phone) = '')
            ORDER BY updated_at ASC NULLS FIRST
            LIMIT %s
        """, [limit])
    else:
        rows = q("""
            SELECT id, name, country FROM bidders
            ORDER BY updated_at ASC NULLS FIRST
            LIMIT %s
        """, [limit])

    results = []
    for row in rows:
        try:
            found = search_company_contact(row["name"], row.get("country") or "")
        except Exception as e:
            results.append(EnrichResult(
                status="error", bidder_id=row["id"],
                bidder_name=row["name"], error=str(e)
            ).dict())
            continue

        if not found:
            results.append(EnrichResult(
                status="no_results", bidder_id=row["id"],
                bidder_name=row["name"]
            ).dict())
            continue

        update_fields = {}
        for key in ("contact_email", "contact_phone", "linkedin_url"):
            val = found.get(key)
            if val:
                update_fields[key] = val

        if update_fields:
            sets = ", ".join(f"{k} = %s" for k in update_fields)
            params = list(update_fields.values()) + [row["id"]]
            with db() as conn:
                with conn.cursor() as cur:
                    cur.execute(f"UPDATE bidders SET {sets}, updated_at = NOW() WHERE id = %s", params)
                conn.commit()

        results.append(EnrichResult(
            status="ok", bidder_id=row["id"],
            bidder_name=row["name"],
            found=found,
            updated_fields=list(update_fields.keys())
        ).dict())

    enriched_count = sum(1 for r in results if r["status"] == "ok")
    return {
        "total_processed": len(results),
        "enriched": enriched_count,
        "results": results,
    }


# ── Fetch country status ──────────────────────────────────────────────────────

@app.get("/api/fetch/countries")
def get_country_fetch_status():
    return get_country_fetch_status_rows()


@app.post("/api/fetch/backfill/{name}")
def trigger_country_backfill(name: str, since: Optional[date] = Query(None)):
    if _fetch_status["running"]:
        return {"status": "already_running", "message": "Another fetch is already in progress"}

    ensure_support_tables()
    country_rows = q("SELECT name FROM target_countries WHERE name = %s", [name])
    if not country_rows:
        raise HTTPException(404, f"{name} is not in target_countries")

    settings = get_app_settings_map()
    if since is None:
        try:
            since = datetime.strptime(settings.get("baseline_date", "2025-01-01"), "%Y-%m-%d").date()
        except ValueError:
            since = date(2025, 1, 1)

    _fetch_status["running"]        = True
    _fetch_status["last_triggered"] = datetime.utcnow().isoformat()
    _fetch_status["last_finished"]  = None
    _fetch_status["last_result"]    = None

    def run_backfill():
        try:
            import fetcher
            fetcher.init_db()
            conn = fetcher.get_connection()
            try:
                upserted, new_records = fetcher.fetch_batch_resilient(conn, [name], since)
                rows = q("SELECT status, error_msg FROM country_fetch_status WHERE country = %s", [name])
                country_status = rows[0]["status"] if rows else "unknown"
                error_msg = rows[0]["error_msg"] if rows else None
                success = country_status != "failed"
                fetcher.log_run(conn, name, upserted, new_records, success, error_msg)
            finally:
                conn.close()

            _fetch_status["last_result"] = {
                "exit_code": 0 if success else 1,
                "stdout": f"{name} backfill finished from {since}: {upserted} fetched, {new_records} new, status={country_status}",
                "stderr": error_msg or "",
                "success": success,
            }
        except Exception as e:
            _fetch_status["last_result"] = {
                "exit_code": -1,
                "stdout": "",
                "stderr": str(e),
                "success": False,
            }
        finally:
            _fetch_status["running"]       = False
            _fetch_status["last_finished"] = datetime.utcnow().isoformat()

    threading.Thread(target=run_backfill, daemon=True).start()
    return {"status": "started", "message": f"{name} backfill started", "since": str(since)}


# ── Other ─────────────────────────────────────────────────────────────────────

@app.get("/api/runs")
def get_runs(limit: int = Query(50)):
    rows = q("SELECT * FROM fetch_runs ORDER BY run_at DESC LIMIT %s", [limit])
    return [dict(r) for r in rows]


@app.get("/health")
def health():
    return {"status": "ok"}
    # echo ("Done")
